from openpyxl import Workbook, load_workbook
from datetime import *
import pickle
from openpyxl.utils import coordinate_to_tuple, get_column_letter


def aquire(sheet,start,step=1,axis ="vertical"):
  coord = coordinate_to_tuple(start)
  row = coord[0]
  col = coord[1]

  step_col = 0
  step_row = 0

  if axis=="horizontal":
    step_col = step
  elif axis == "vertical":
    step_row = step
  else:
    print("Axis argument should be either 'vertical' or 'horizontal'")
    return NoneType
  datas = []
  while True:
    data = sheet[get_column_letter(col) +str(row)].value
    if not data:
      break
    datas.append(data)
    row += step_row
    col += step_col

  return datas





def generate_tree_info(animais_data_peso,animais_tree_info):
  animais_tree_info = []
  for animal in animais_data_peso:
    animais_tree_info.append([animal, 0, datetime(1900, 1, 1, 0, 0), 0, "Não"])

  animais_tree_info.sort()

  for i, animal in enumerate(animais_tree_info):
    animal[3] = len(animais_data_peso[animal[0]])

  for i, animal in enumerate(animais_tree_info):
    try:
      animal[1] = animais_data_peso[animal[0]][0][1]
      animal[2] = animais_data_peso[animal[0]][0][0]
    except IndexError:
      pass

  return animais_tree_info




if __name__ == "__main__":
  workbook = load_workbook(filename="gado_py.xlsx", data_only=True)
  name_sheet = workbook.sheetnames[1]
  sheet = workbook[name_sheet]

  col_ani = coordinate_to_tuple("A12")[1]
  row = 12
  all_animais = aquire(sheet, start="A12")
  all_datas = aquire(sheet, start="E10", step=2, axis="horizontal")

  animais_data_peso = {}

  for i in range(len(all_animais)):
    tem_pesagem = False
    col_data = coordinate_to_tuple("E12")[1]

    for j in range(len(all_datas)):
      peso = sheet[get_column_letter(col_data) + str(row)].value
      col_data += 2
      if peso and not isinstance(peso, str):
        if not tem_pesagem:
          animais_data_peso[all_animais[i]] = []
          tem_pesagem = True
        animais_data_peso[all_animais[i]].append((all_datas[j], peso))
      if tem_pesagem:
        animais_data_peso[all_animais[i]] = sorted(animais_data_peso[all_animais[i]], reverse=True)
    row += 1


  animais_tree_info = generate_tree_info(animais_data_peso)

  file = open("animais_data_peso.pkl","wb")
  pickle.dump(animais_data_peso,file)
  file.close()

  file = open("all_animais.pkl","wb")
  pickle.dump(all_animais,file)
  file.close()

  file = open("all_datas.pkl","wb")
  pickle.dump(all_datas,file)
  file.close()

  file = open("animais_tree_info.pkl","wb")
  pickle.dump(animais_tree_info,file)
  file.close()